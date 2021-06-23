###############################################################
################# https://www.fardanesh.ir ####################
###############################################################


from openpyxl import load_workbook

from openpyxl.styles import PatternFill, GradientFill 



wb=load_workbook('lecture06-styles/list1.xlsx')
ws1=wb.active


fill1=PatternFill("solid", fgColor="DDDDDD")
fill2 = GradientFill(stop=("000000", "FFFFFF"))



for row in ws1['A1:H1']:
    for c in row:
        c.fill=fill1

for row in ws1['A3:H3']:
    for c in row:
        c.fill=fill2


wb.save(filename='lecture06-styles/list1Modified.xlsx')
