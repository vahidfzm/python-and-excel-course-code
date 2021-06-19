###############################################################
################# https://www.fardanesh.ir ####################
###############################################################


from openpyxl import Workbook


wb=Workbook()
ws1 = wb.active

ws1['A1'].value=33
ws1['A2']=22
ws1['A3']="=sum(A1:A2)"
ws1['A4']="Fardanesh.ir"


# this is not correct
# ws1['B1:B4']="Fardanesh.ir"
# print(ws1['B1:B4'])


# set value in the cell B2
cellB2=ws1.cell(row=2,column=2)
cellB2.value=66

# set value in the cell B3
ws1.cell(row=3,column=2,value=99)


# fill numbers 1..10 in range C1:C10
for i in range(10):
    ws1.cell(row=i+1,column=3,value=i+1)

wb.save(filename='lecture02/file1.xlsx')
