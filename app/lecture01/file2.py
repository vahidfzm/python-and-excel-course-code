###############################################################
################# https://www.fardanesh.ir ####################
###############################################################


from openpyxl import load_workbook


# open a workbook
wb=load_workbook('lecture01/list1.xlsx')

# add a new sheet at the end (default behavior)
ws1 = wb.create_sheet("MysheetAtTheEnd") 

# add a new sheet at first position
ws2 = wb.create_sheet("MysheetAtTheBeginning", 0)

# remove a sheet
wb.remove(wb['Sheet3'])

# copy a worksheet
cpws=wb.copy_worksheet(wb['Sheet1'])
cpws.title="My copy from Sheet1"


# save the workbook
wb.save(filename='lecture01/list1Copy.xlsx')
