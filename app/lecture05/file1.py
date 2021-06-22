###############################################################
################# https://www.fardanesh.ir ####################
###############################################################


from openpyxl import load_workbook

wb=load_workbook('lecture05/list1.xlsx')

ws1=wb.active


# https://openpyxl.readthedocs.io/en/stable/_modules/index.html



# add a new row before the existing row 5:
# ws1.insert_rows(5)

# add 10 new rows before the existing row 5:
# ws1.insert_rows(5,amount=10)
# ws1.insert_rows(5,10)


# add a new column before the existing column 1:
# ws1.insert_cols(1)


# add 3 new column before the existing column 1:
# ws1.insert_cols(5,amount=3)

# delete row 2
# ws1.delete_rows(2)


# delete rows 2 to 10
# ws1.delete_rows(2,8)


# move a range without any change in formulas
# ws1.move_range("A1:H20", rows=6, cols=2)


# move a range wit formula-translation
# ws1.move_range("A1:H20", rows=6, cols=2,translate=True)


# merge cells
# ws1.merge_cells('A1:A5')
# ws1.merge_cells(start_row=2, start_column=2, end_row=8, end_column=8)

# ws1.unmerge_cells('A1:A5')


# ws1.freeze_panes='D2'
# ws1.freeze_panes=None


wb.save(filename='lecture05/list1Modified.xlsx')
